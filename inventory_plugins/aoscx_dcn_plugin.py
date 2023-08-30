#!/usr/bin/env python
# aoscx_dcn_plugin.py

from __future__ import (absolute_import, division, print_function)
__metaclass__ = type

DOCUMENTATION = r'''
    name: aoscx_dcn_plugin
    plugin_type: inventory
    short_description: Returns Ansible CX Data Center inventory from Excel
    description: Returns Ansible CX Data Center inventory from Excel
    options:
      plugin:
          description: Name of the plugin
          required: true
          choices: ['aoscx_dcn_plugin']
      path_to_inventory:
        description: Full path of the directory containing Excel inventory file
        required: true
      excel_file:
        description: File name of the Excel inventory file
        required: true
'''

from ansible.plugins.inventory import BaseInventoryPlugin
from ansible.errors import AnsibleError, AnsibleParserError
from openpyxl import load_workbook
from netaddr import IPNetwork, AddrFormatError
import sys


class InventoryModule(BaseInventoryPlugin):
    NAME = 'aoscx_dcn_plugin'

    ARCH1_EXCEL_SHEETS = ['Network', 'DC ToR Network']
    ARCH2_EXCEL_SHEETS = ['Network', 'Core Network', 'Access Network']
    ARCH3_EXCEL_SHEETS = ['Network', 'Spine Network', 'Leaf Network']

    COMMON_EXCEL_DICT = {'Device Hostname': 'hostname',
                         'Switch OOBM IP': 'ansible_host',
                         'Loopback Interface 0 IP Address': 'loopback0_ip',
                         'VLAN ID': 'vlan_id',
                         'Description': 'description',
                         'Active Gateway IP': 'active_gateway_ip',
                         'Active Gateway MAC': 'active_gateway_mac',
                         'Jinja2 Config Template': 'config_template',
                         'Switch Login Username': 'ansible_user',
                         'Switch Login Password': 'ansible_password',
                         'Validate Switch SSL Certificate':
                         'ansible_httpapi_validate_certs',
                         'Generated Configuration Destination Path':
                         'config_path',
                         'MTU': 'mtu',
                         'Loopback Addresses': 'loopback_subnet',
                         '10g Speed Interface Group':
                         'speed_interface_group_10g'
                         }

    VSX_EXCEL_DICT = {'VSX Role': 'vsx_role',
                      'VSX Primary Keepalive IP Address':
                      'vsx_keepalive_ip_primary',
                      'VSX Secondary Keepalive IP Address':
                      'vsx_keepalive_ip_secondary',
                      'VSX System Mac Address': 'vsx_system_mac',
                      'VSX Keepalive Interface': 'vsx_keepalive_int',
                      'VSX ISL Ports': 'vsx_isl_ports',
                      'VSX ISL Lag ID': 'vsx_isl_lagid'
                      }

    ARCH1_EXCEL_DICT = {'iBGP Transit VLAN IP Address': 'ibgp_transit_vlan_ip',
                        'Peer1 VLAN IP': 'peer1_vlan_ipv4',
                        'Peer2 VLAN IP': 'peer2_vlan_ipv4',
                        'OSPF Area': 'ospf_area',
                        'BGP ASN': 'asn',
                        'STP Priority': 'stp_priority',
                        'STP Revision Number': 'stp_rev_num',
                        'STP Configuration Name': 'stp_config_name',
                        'iBGP Transit VLAN': 'ibgp_transit_vlan',
                        'MCLAG Downlink Interfaces':
                        'access_mclag_downlink_ports',
                        'MCLAG ID': 'mclag_id',
                        'MCLAG Description': 'mclag_description',
                        'MCLAG Trunk VLANs': 'trunk_vlans'
                        }

    ARCH2_EXCEL_DICT = {'iBGP Transit VLAN IP Address': 'ibgp_transit_vlan_ip',
                        'Access VSX Pair': 'vsx_pair',
                        '10g Speed Interface Group':
                        'speed_interface_group_10g',
                        'Core Uplink MCLAG ID': 'core_mclag_id',
                        'Core Uplink MCLAG Ports': 'core_mclag_uplink_ports',
                        'Core Uplink MCLAG Description':
                        'core_mclag_description',
                        'Server VLANs': 'server_vlans',
                        'OSPF Area': 'ospf_area',
                        'BGP ASN': 'asn',
                        'STP Priority': 'stp_priority',
                        'STP Revision Number': 'stp_rev_num',
                        'STP Configuration Name': 'stp_config_name',
                        'iBGP Transit VLAN': 'ibgp_transit_vlan',
                        'MCLAG Downlink Interfaces':
                        'access_mclag_downlink_ports',
                        'MCLAG ID': 'mclag_id',
                        'MCLAG Description': 'mclag_description',
                        'MCLAG Trunk VLANs': 'trunk_vlans'
                        }

    ARCH3_EXCEL_DICT = {'BGP ASN': 'asn',
                        'Fabric Addresses': 'fabric_subnet',
                        'Leaf VSX Pair Downlink Interfaces':
                        'vsx_pair_downlink_interfaces',
                        'Leaf VSX Pair': 'vsx_pair',
                        'Layer3 Fabric Design (iBGP / eBGP)': 'fabric',
                        'Loopback Interface 1 IP Address': 'loopback1_ip',
                        'Spine Uplink Ports': 'spine_uplink_ports',
                        'Server VLANs': 'server_vlans'
                        }

    ARCH3_EBGP_EXCEL_DICT = {'Leaf BGP ASN': 'vsx_pair_asn'}

    ARCH3_IBGP_EXCEL_DICT = {'OSPF Area': 'ospf_area'
                             }

    def _is_string(self, tmp_var):
        # Check if variable is bytes (python 3) or unicode (python 2) or str
        # Compatible between python2 and python3 versions
        if str(type(tmp_var)) in ["<class 'bytes'>", "<type 'unicode'>",
                                  "<type 'str'>", "<class 'str'>"]:
            return True
        else:
            return False

    def _generate_loopback(self, excel_data, architecture):
        """
        Generates loopback addresses for interface Loopback 0 and interface
        Loopback 1 for all hosts when Loopback Addresses is provded in excel

        :param excel_data (dict): a dictionary containing processed information
        from excel file
        :param architecture (str): a string describing which DCN architecture
        is being processed

        Returns:
            excel_data: a dictionary containing newly generated loopback
            addresses for each host and processed information from
            excel file
        """

        if 'loopback_subnet' not in excel_data['network'].keys():
            return excel_data

        loopback_network = excel_data['network']['loopback_subnet']

        if '/' not in loopback_network:
            loopback_network = loopback_network + "/24"

        try:
            loopback_network = IPNetwork(loopback_network)

        except AddrFormatError:
            sys.exit("Error with Loopback Addresses "
                     "value : {tmp_val}".format(tmp_val=loopback_network))

        loopback_subnets = list(loopback_network.subnet(32))
        subnet_counter = 1
        host_data = excel_data['hosts']

        if architecture == 'ARCH1':
            for device in sorted(excel_data['dc_tor']):
                formatted_ip = str(loopback_subnets[subnet_counter + 10].ip)
                host_data[device]['loopback0_ip'] = formatted_ip
                subnet_counter += 1
        elif architecture == 'ARCH2':
            for device in sorted(excel_data['core']):
                formatted_ip = str(loopback_subnets[subnet_counter + 10].ip)
                host_data[device]['loopback0_ip'] = formatted_ip
                subnet_counter += 1
            for device in sorted(excel_data['access']):
                formatted_ip = str(loopback_subnets[subnet_counter].ip)
                host_data[device]['loopback0_ip'] = formatted_ip
                subnet_counter += 1
        elif 'ARCH3' in architecture:
            for device in sorted(excel_data['spine']):
                formatted_ip = str(loopback_subnets[subnet_counter + 10].ip)
                host_data[device]['loopback0_ip'] = formatted_ip
                subnet_counter += 1
            for device in sorted(excel_data['leaf']):
                formatted_ip0 = str(loopback_subnets[subnet_counter].ip)
                host_data[device]['loopback0_ip'] = formatted_ip0
                formatted_ip1 = str(loopback_subnets[subnet_counter + 100].ip)
                host_data[device]['loopback1_ip'] = formatted_ip1
                subnet_counter += 1

        return excel_data

    def _convert_network_sheet(self, excel_data, network_sheet, architecture):
        """
        Converts the Network sheet for every excel workbook

        :param excel_data (dict): a dictionary containing processed information
        from excel file
        :param network_sheet (openpyxl sheet): openpyxl Network sheet from
        workbook
        :param architecture (str): a string describing which DCN architecture
        is being processed

        Returns:
            excel_data: a dictionary containing newly processed Network data
            previously and processed information from excel file
        """
        data_row = -1
        network_data = {}

        for row in network_sheet.iter_rows():
            if row[0].row == 1:
                continue
            elif row[0].value == 'Management Network Details':
                data_row = row[0].row + 1
            # Process data for extra variables under Network section
            elif data_row <= row[0].row and row[0].value:
                if self._is_string(row[1].value):
                    if ',' in row[1].value:
                        tmp_var = row[1].value.split(',')
                        tmp_var = list(filter(None, tmp_var))
                        tmp_list = [x.strip() for x in tmp_var]
                        network_data[row[0].value] = tmp_list
                        continue
                    row[1].value = row[1].value.strip()
                network_data[row[0].value] = row[1].value

        if architecture == 'ARCH1':
            arch_value_dict = self.ARCH1_EXCEL_DICT
        elif architecture == 'ARCH2':
            arch_value_dict = self.ARCH2_EXCEL_DICT
        elif architecture == 'ARCH3':
            arch_value_dict = self.ARCH3_EXCEL_DICT

        converted_network_data = {}
        for tmp_key, tmp_value in network_data.items():

            if tmp_key in arch_value_dict.keys():
                converted_network_data[arch_value_dict[tmp_key]] = tmp_value
                if tmp_key == 'Layer3 Fabric Design (iBGP / eBGP)' and tmp_value in ['iBGP', 'eBGP']:
                    excel_data['architecture'] = "ARCH3_{fabric}".format(fabric=tmp_value.upper())
            elif tmp_key in self.COMMON_EXCEL_DICT.keys():
                converted_network_data[self.COMMON_EXCEL_DICT[tmp_key]] = tmp_value
            elif tmp_key in self.VSX_EXCEL_DICT.keys():
                converted_network_data[self.VSX_EXCEL_DICT[tmp_key]] = tmp_value
            elif architecture == 'ARCH3':
                if tmp_key in self.ARCH3_EBGP_EXCEL_DICT.keys():
                    converted_network_data[self.ARCH3_EBGP_EXCEL_DICT[tmp_key]] = tmp_value
                elif tmp_key in self.ARCH3_IBGP_EXCEL_DICT.keys():
                    converted_network_data[self.ARCH3_IBGP_EXCEL_DICT[tmp_key]] = tmp_value
            else:
                converted_network_data[tmp_key] = tmp_value

        excel_data['network'] = converted_network_data
        return excel_data

    def _convert_core_sheet(self, excel_data, core_sheet, architecture):
        """
        Converts the Core Network sheet for the Dedicated Two-Tier VRD excel
        workbook

        :param excel_data (dict): a dictionary containing processed information
        from excel file
        :param core_sheet (openpyxl sheet): openpyxl Core Network sheet from
        workbook
        :param architecture (str): a string describing which DCN architecture
        is being processed

        Returns:
            excel_data: a dictionary containing newly processed Core data
            previously and processed information from excel file
        """
        data_row = -1
        core_data = {}
        network_details = False
        vsx_pair_info = False
        core_svi_info = False
        common_values = False
        core_svi_headers = []
        vsx_pair_number_row = None
        vsx_pair_info_dict = {}
        peer1, peer2 = '', ''

        if architecture == 'ARCH1':
            arch_value_dict = self.ARCH1_EXCEL_DICT
        elif architecture == 'ARCH2':
            arch_value_dict = self.ARCH2_EXCEL_DICT

        for row in core_sheet.iter_rows():
            if row[0].row == 1:
                continue
            # Retreiving host data from section starting with 'Network Details'
            elif row[0].value == 'Network Details':
                network_details = True
                data_row = row[0].row + 1
                # Retrieves headers from next row
                tmp_row = core_sheet[row[0].row + 1]
                if tmp_row[0].value == 'Device Hostname':
                    peer1 = tmp_row[1].value
                    peer2 = tmp_row[2].value
                    if architecture == 'ARCH1':
                        excel_data['dc_tor'].append(peer1)
                        excel_data['dc_tor'].append(peer2)
                        core_data[peer1] = {'role': 'dc_tor'}
                        core_data[peer2] = {'role': 'dc_tor'}
                    elif architecture == 'ARCH2':
                        excel_data['core'].append(peer1)
                        excel_data['core'].append(peer2)
                        core_data[peer1] = {'role': 'core'}
                        core_data[peer2] = {'role': 'core'}
            elif row[0].value == 'Core Common Values' or row[0].value == 'DC ToR Common Values':
                common_values = True
                data_row = row[0].row + 2
            elif row[0].value == 'Access VSX Pair Information' or 'Server MCLAG Information' == row[0].value and not vsx_pair_info:
                vsx_pair_info = True
                data_row = row[0].row + 2
                tmp_row = core_sheet[row[0].row + 1]
                if tmp_row[0].value == 'Access VSX Pair' or 'Server MCLAG Information' == row[0].value:
                    vsx_pair_number_row = tmp_row[0].row
                    # Retrieves hostnames from next row
                    for tmp in tmp_row[1:]:
                        if tmp.value:
                            vsx_pair_info_dict[tmp.value.lower()] = {}

            elif row[0].value == 'Core SVI Information' or row[0].value == 'DC ToR SVI Information':
                core_svi_info = True
                data_row = row[0].row + 2
                for tmp in core_sheet[row[0].row + 1]:
                    if tmp.value:
                        core_svi_headers.append(tmp.value)
            elif not row[0].value and (network_details or core_svi_info or common_values or vsx_pair_info):
                if network_details:
                    network_details = False
                elif vsx_pair_info:
                    vsx_pair_info = False
                    for vsx_pair, vsx_pair_data in vsx_pair_info_dict.items():
                        vsx_pair = vsx_pair.lower()
                        for core in core_data.keys():
                            if 'vsx_pair_mclags' not in core_data[core].keys():
                                core_data[core]['vsx_pair_mclags'] = []
                            vsx_pair_info_dict[vsx_pair].update({'location': vsx_pair})
                            core_data[core]['vsx_pair_mclags'].append(vsx_pair_info_dict[vsx_pair])

                elif core_svi_info:
                    core_svi_info = False
                elif common_values:
                    common_values = False
                data_row = -1
            # Process data for each VLAN under Core SVI Information or DC ToR SVI Information
            elif data_row <= row[0].row and row[0].value and core_svi_info:
                tmp_dict = {}
                # Process data for each cell in the row
                for tmp in row:
                    if tmp.value:
                        # Ignore any cell where value is empty or n/a
                        if tmp.value in ['n/a', 'N/A', None, ""]:
                            continue
                        elif self._is_string(tmp.value):
                            # Strip any extra white space
                            tmp.value = tmp.value.strip()
                        tmp_key = core_svi_headers[row.index(tmp)]
                        if tmp_key in arch_value_dict.keys():
                            tmp_key = arch_value_dict[tmp_key]
                        elif tmp_key in self.COMMON_EXCEL_DICT.keys():
                            tmp_key = self.COMMON_EXCEL_DICT[tmp_key]
                        tmp_dict[tmp_key] = tmp.value
                if 'Peer1 VLAN IP / MASK' in tmp_dict.keys() and 'Peer2 VLAN IP / MASK' in tmp_dict.keys():
                    tmp_peer1_ip = tmp_dict.pop('Peer1 VLAN IP / MASK')
                    tmp_peer2_ip = tmp_dict.pop('Peer2 VLAN IP / MASK')

                    if 'server_vlans' not in core_data[peer1].keys():
                        core_data[peer1]['server_vlans'] = []
                    core_data[peer1]['server_vlans'].append(tmp_dict['vlan_id'])
                    if 'server_vlans' not in core_data[peer2].keys():
                        core_data[peer2]['server_vlans'] = []
                    core_data[peer2]['server_vlans'].append(tmp_dict['vlan_id'])

                    if 'core_vlan_interfaces' not in core_data[peer1].keys():
                        core_data[peer1]['core_vlan_interfaces'] = []
                    tmp_dict.update({'ipv4': tmp_peer1_ip})
                    core_data[peer1]['core_vlan_interfaces'].append(tmp_dict.copy())
                    tmp_dict.pop('ipv4')
                    if 'core_vlan_interfaces' not in core_data[peer2].keys():
                        core_data[peer2]['core_vlan_interfaces'] = []
                    tmp_dict.update({'ipv4': tmp_peer2_ip})
                    core_data[peer2]['core_vlan_interfaces'].append(tmp_dict.copy())

            # Process data for each peer under Access Rack information
            elif data_row <= row[0].row and row[0].value and vsx_pair_info:
                tmp_key = row[0].value
                if tmp_key in arch_value_dict.keys():
                    tmp_key = arch_value_dict[tmp_key]
                elif tmp_key in self.COMMON_EXCEL_DICT.keys():
                    tmp_key = self.COMMON_EXCEL_DICT[tmp_key]
                elif tmp_key in self.VSX_EXCEL_DICT.keys():
                    tmp_key = self.VSX_EXCEL_DICT[tmp_key]

                # Convert matching cells into list values
                for tmp in row:
                    tmp_value = tmp.value
                    # Ignore any cell where value is empty or n/a
                    if tmp_value in ['n/a', 'N/A', None, ""]:
                        continue
                    if tmp.column is 1:
                        continue
                    if tmp_value:
                        column_key = core_sheet[vsx_pair_number_row][tmp.column-1].value.lower()
                        if self._is_string(tmp_value):
                            if ',' in tmp_value:
                                tmp_var = tmp_value.split(',')
                                tmp_var = list(filter(None, tmp_var))
                                tmp_list = [x.strip() for x in tmp_var]
                                if column_key:
                                    vsx_pair_info_dict[column_key][tmp_key] = tmp_list
                                continue

                            tmp_value = tmp_value.strip()

                        if tmp_key in ['trunk_vlans', 'vsx_isl_ports']:
                            if not isinstance(tmp_value, list):
                                tmp_value = [tmp_value]

                        if column_key:
                            vsx_pair_info_dict[column_key][tmp_key] = tmp_value

            # Process data for each peer under Network Details or Common Core Values
            elif data_row <= row[0].row and row[0].value and (network_details or common_values):
                tmp_key = row[0].value
                if tmp_key in arch_value_dict.keys():
                    tmp_key = arch_value_dict[tmp_key]
                elif tmp_key in self.COMMON_EXCEL_DICT.keys():
                    tmp_key = self.COMMON_EXCEL_DICT[tmp_key]
                elif tmp_key in self.VSX_EXCEL_DICT.keys():
                    tmp_key = self.VSX_EXCEL_DICT[tmp_key]

                # Convert matching cells into list values
                for i in range(1, 3):
                    tmp_value = row[i].value
                    # Ignore any cell where value is empty or n/a
                    if tmp_value in ['n/a', 'N/A', None, ""]:
                        continue
                    if self._is_string(tmp_value):
                        if ',' in tmp_value:
                            tmp_var = tmp_value.split(',')
                            tmp_var = list(filter(None, tmp_var))
                            tmp_list = [x.strip() for x in tmp_var]
                            if network_details:
                                if i == 1:
                                    core_data[peer1][tmp_key] = tmp_list
                                else:
                                    core_data[peer2][tmp_key] = tmp_list
                            elif common_values:
                                core_data[peer1][tmp_key] = tmp_list
                                core_data[peer2][tmp_key] = tmp_list
                            continue
                        row[i].value = row[i].value.strip()

                if network_details:
                    core_data[peer1][tmp_key] = row[1].value
                    core_data[peer2][tmp_key] = row[2].value
                elif common_values:
                    core_data[peer1][tmp_key] = row[1].value
                    core_data[peer2][tmp_key] = row[1].value

        for core in core_data.keys():
            for tmp_key, tmp_value in core_data[core].items():
                if tmp_key in ['server_vlans', 'trunk_vlans', 'vsx_isl_ports']:
                        if not isinstance(tmp_value, list):
                            core_data[core][tmp_key] = [tmp_value]

        if architecture == 'ARCH2':
            core_data[peer1]['role'], core_data[peer2]['role'] = 'core', 'core'
        elif architecture == 'ARCH1':
            core_data[peer1]['role'], core_data[peer2]['role'] = 'dc_tor', 'dc_tor'

        if core_data[peer1]['vsx_role'] == 'primary':
            core_data[peer1]['core_vsx_neighbor_ip'] = core_data[peer1]['vsx_keepalive_ip_secondary']
        else:
            core_data[peer1]['core_vsx_neighbor_ip'] = core_data[peer1]['vsx_keepalive_ip_primary']

        if core_data[peer2]['vsx_role'] == 'primary':
            core_data[peer2]['core_vsx_neighbor_ip'] = core_data[peer2]['vsx_keepalive_ip_secondary']
        else:
            core_data[peer2]['core_vsx_neighbor_ip'] = core_data[peer2]['vsx_keepalive_ip_primary']

        for device, device_data in core_data.items():
            tmp_vlans = list(str(item) for item in device_data['server_vlans'])
            for mclag in device_data['vsx_pair_mclags']:
                tmp_trunk_vlans = list(str(item)
                                       for item in mclag['trunk_vlans'])
                if not all(item in tmp_vlans for item in tmp_trunk_vlans):
                    sys.exit('Error in Core/DC ToR VLAN data in Core/DC ToR'
                             'sheet - MCLAG Trunk VLANs {0} must exist in '
                             'Core/DC ToR SVI Information {1}'
                             ''.format(tmp_trunk_vlans, tmp_vlans))

        excel_data['hosts'].update(core_data)

        return excel_data

    def _convert_access_sheet(self, excel_data, access_sheet):
        """
        Converts the Access Network sheet for the Dedicated Two-Tier VRD excel
        workbook

        :param excel_data (dict): a dictionary containing processed information
        from excel file
        :param access_sheet (openpyxl sheet): openpyxl Access Network sheet from
        workbook

        Returns:
            excel_data: a dictionary containing newly processed Access data
            previously and processed information from excel file
        """
        data_row = -1
        access_data = {}
        network_details = False
        vsx_pair_info = False
        common_values = False
        access_hostname_row = None
        vsx_pair_number_row = None
        vsx_pair_info_dict = {}

        for row in access_sheet.iter_rows():
            if row[0].row == 1:
                continue
            # Retreiving host data from section starting with 'Network Details'
            elif row[0].value == 'Network Details':
                network_details = True
                data_row = row[0].row + 1

                tmp_row = access_sheet[row[0].row + 1]
                if tmp_row[0].value == 'Device Hostname':
                    access_hostname_row = tmp_row[0].row
                    # Retrieves hostnames from next row
                    for tmp in tmp_row[1:]:
                        if tmp.value:
                            access_data[tmp.value] = {'role': 'access'}

            elif row[0].value == 'Access Common Values':
                common_values = True
                data_row = row[0].row + 2
            elif row[0].value == 'Access VSX Pair Information':
                vsx_pair_info = True
                data_row = row[0].row + 2

                tmp_row = access_sheet[row[0].row + 1]
                if tmp_row[0].value == 'Access VSX Pair':
                    vsx_pair_number_row = tmp_row[0].row
                    # Retrieves hostnames from next row
                    for tmp in tmp_row[1:]:
                        if tmp.value:
                            vsx_pair_info_dict[tmp.value.lower()] = {}

            elif not row[0].value and (network_details or vsx_pair_info or common_values):
                if network_details:
                    network_details = False
                elif vsx_pair_info:
                    vsx_pair_info = False
                    for device, device_data in access_data.items():
                        if 'vsx_pair' in device_data.keys():
                            access_data[device].update(vsx_pair_info_dict[device_data['vsx_pair']])
                elif common_values:
                    common_values = False
                data_row = -1
            # Process data for Access Common Values
            elif data_row <= row[0].row and row[0].value and common_values:
                tmp_key = row[0].value
                tmp_value = row[1].value

                if tmp_value and tmp_key:
                    # Ignore any cell where value is empty or n/a
                    if tmp_value in ['n/a', 'N/A', None, ""]:
                        continue
                    elif self._is_string(tmp_value):
                        if ',' in tmp_value:
                            tmp_var = tmp_value.split(',')
                            tmp_var = list(filter(None, tmp_var))
                            tmp_list = [x.strip() for x in tmp_var]

                            for device in access_data.keys():
                                access_data[device][tmp_key] = tmp_list

                        # Strip any extra white space
                        tmp_value = tmp_value.strip()

                    if tmp_key in self.ARCH2_EXCEL_DICT.keys():
                        tmp_key = self.ARCH2_EXCEL_DICT[tmp_key]
                    elif tmp_key in self.COMMON_EXCEL_DICT.keys():
                        tmp_key = self.COMMON_EXCEL_DICT[tmp_key]
                    elif tmp_key in self.VSX_EXCEL_DICT.keys():
                        tmp_key = self.VSX_EXCEL_DICT[tmp_key]

                    if tmp_key in ['server_vlans', 'trunk_vlans', 'vsx_isl_ports']:
                            if not isinstance(tmp_value, list):
                                tmp_value = [tmp_value]

                    if tmp_key == 'vsx_pair':
                        tmp_value = tmp_value.lower()

                    for device in access_data.keys():
                        access_data[device][tmp_key] = tmp_value

            # Process data for each peer under Network Details or Access Rack information
            elif data_row <= row[0].row and row[0].value and (network_details or vsx_pair_info):
                tmp_key = row[0].value
                if tmp_key in self.ARCH2_EXCEL_DICT.keys():
                    tmp_key = self.ARCH2_EXCEL_DICT[tmp_key]
                elif tmp_key in self.COMMON_EXCEL_DICT.keys():
                    tmp_key = self.COMMON_EXCEL_DICT[tmp_key]
                elif tmp_key in self.VSX_EXCEL_DICT.keys():
                    tmp_key = self.VSX_EXCEL_DICT[tmp_key]

                # Convert matching cells into list values
                for tmp in row:
                    tmp_value = tmp.value
                    # Ignore any cell where value is empty or n/a
                    if tmp_value in ['n/a', 'N/A', None, ""]:
                        continue
                    if tmp.column is 1:
                        continue
                    if tmp_value:
                        if network_details:
                            column_key = access_sheet[access_hostname_row][tmp.column-1].value
                        elif vsx_pair_info:
                            column_key = access_sheet[vsx_pair_number_row][tmp.column-1].value.lower()
                            tmp_key = tmp_key.lower()
                        if self._is_string(tmp_value):
                            if ',' in tmp_value:
                                tmp_var = tmp_value.split(',')
                                tmp_var = list(filter(None, tmp_var))
                                tmp_list = [x.strip() for x in tmp_var]
                                if column_key and network_details:
                                    access_data[column_key][tmp_key] = tmp_list
                                elif column_key and vsx_pair_info:
                                    vsx_pair_info_dict[column_key][tmp_key] = tmp_list
                                continue

                            tmp_value = tmp_value.strip()

                        if tmp_key in ['server_vlans', 'vsx_isl_ports']:
                            if not isinstance(tmp_value, list):
                                tmp_value = [tmp_value]

                        if tmp_key == 'vsx_pair':
                            tmp_value = tmp_value.lower()

                        if column_key and network_details:
                            access_data[column_key][tmp_key] = tmp_value
                        elif column_key and vsx_pair_info:
                            vsx_pair_info_dict[column_key][tmp_key] = tmp_value

        for access in access_data.keys():
            excel_data['access'].append(access)

        excel_data['hosts'].update(access_data)

        return excel_data

    def _convert_spine_sheet(self, excel_data, spine_sheet, architecture):
        """
        Converts the Spine Network sheet for the Spine/Leaf VRD excel workbook

        :param excel_data (dict): a dictionary containing processed information
        from excel file
        :param spine_sheet (openpyxl sheet): openpyxl Spine Network sheet from
        workbook
        :param architecture (str): a string describing which DCN architecture
        is being processed

        Returns:
            excel_data: a dictionary containing newly processed Spine data
            previously and processed information from excel file
        """
        data_row = -1
        spine_data = {}
        network_details = False
        vsx_pair_info = False
        common_values = False
        vsx_pair_number_row = None
        vsx_pair_info_dict = {}

        if architecture == 'ARCH3_EBGP':
            arch_value_dict = self.ARCH3_EBGP_EXCEL_DICT.copy()
            arch_value_dict.update(self.ARCH3_EXCEL_DICT)
        elif architecture == 'ARCH3_IBGP':
            arch_value_dict = self.ARCH3_IBGP_EXCEL_DICT.copy()
            arch_value_dict.update(self.ARCH3_EXCEL_DICT)

        for row in spine_sheet.iter_rows():
            if row[0].row == 1:
                continue
            # Retreiving host data from section starting with 'Network Details'
            elif row[0].value == 'Network Details':
                network_details = True
                data_row = row[0].row + 1
                # Retrieves headers from next row
                tmp_row = spine_sheet[row[0].row + 1]
                if tmp_row[0].value == 'Device Hostname':
                        peer1 = tmp_row[1].value
                        peer2 = tmp_row[2].value
                        excel_data['spine'].append(peer1)
                        excel_data['spine'].append(peer2)
                        spine_data[peer1] = {'role': 'spine'}
                        spine_data[peer2] = {'role': 'spine'}
            elif row[0].value == 'Spine Common Values':
                common_values = True
                data_row = row[0].row + 2
            elif row[0].value == 'Leaf VSX Pair Information':
                vsx_pair_info = True
                data_row = row[0].row + 2
                tmp_row = spine_sheet[row[0].row + 1]
                if tmp_row[0].value == 'Leaf VSX Pair':
                    vsx_pair_number_row = tmp_row[0].row
                    # Retrieves Leaf VSX Pair Groups from next row
                    for tmp in tmp_row[1:]:
                        if tmp.value:
                            vsx_pair_info_dict[tmp.value.lower()] = {}

            elif not row[0].value and (network_details or vsx_pair_info or common_values):
                if network_details:
                    network_details = False
                elif vsx_pair_info:
                    vsx_pair_info = False

                    for vsx_pair, vsx_pair_data in vsx_pair_info_dict.items():
                        vsx_pair = vsx_pair.lower()
                        for spine in spine_data.keys():
                            if 'vsx_pair_downlinks' not in spine_data[spine].keys():
                                spine_data[spine]['vsx_pair_downlinks'] = {}
                            spine_data[spine]['vsx_pair_downlinks'][vsx_pair] = vsx_pair_data
                elif common_values:
                    common_values = False
                data_row = -1

            # Process data for each peer under Leaf VSX Pair information
            elif data_row <= row[0].row and row[0].value and vsx_pair_info:
                tmp_key = row[0].value
                if tmp_key in arch_value_dict.keys():
                    tmp_key = arch_value_dict[tmp_key]
                elif tmp_key in self.COMMON_EXCEL_DICT.keys():
                    tmp_key = self.COMMON_EXCEL_DICT[tmp_key]
                elif tmp_key in self.VSX_EXCEL_DICT.keys():
                    tmp_key = self.VSX_EXCEL_DICT[tmp_key]

                # Convert matching cells into list values
                for tmp in row:
                    tmp_value = tmp.value
                    # Ignore any cell where value is empty or n/a
                    if tmp_value in ['n/a', 'N/A', None, ""]:
                        continue
                    if tmp.column is 1:
                        continue
                    if tmp_value:
                        column_key = spine_sheet[vsx_pair_number_row][tmp.column-1].value.lower()
                        if self._is_string(tmp_value):
                            if ',' in tmp_value:
                                tmp_var = tmp_value.split(',')
                                tmp_var = list(filter(None, tmp_var))
                                tmp_list = [x.strip() for x in tmp_var]
                                if column_key:
                                    vsx_pair_info_dict[column_key][tmp_key] = tmp_list
                                continue

                            tmp_value = tmp_value.strip()

                        if column_key:
                            vsx_pair_info_dict[column_key][tmp_key] = tmp_value

            # Process data for each peer under Network Details or Common Core Values
            elif data_row <= row[0].row and row[0].value and (network_details or common_values):
                tmp_key = row[0].value
                if tmp_key in arch_value_dict.keys():
                    tmp_key = arch_value_dict[tmp_key]
                elif tmp_key in self.COMMON_EXCEL_DICT.keys():
                    tmp_key = self.COMMON_EXCEL_DICT[tmp_key]
                elif tmp_key in self.VSX_EXCEL_DICT.keys():
                    tmp_key = self.VSX_EXCEL_DICT[tmp_key]

                # Convert matching cells into list values
                for i in range(1, 3):
                    tmp_value = row[i].value
                    # Ignore any cell where value is empty or n/a
                    if tmp_value in ['n/a', 'N/A', None, ""]:
                        continue
                    if self._is_string(tmp_value):
                        if ',' in tmp_value:
                            tmp_var = tmp_value.split(',')
                            tmp_var = list(filter(None, tmp_var))
                            tmp_list = [x.strip() for x in tmp_var]
                            if network_details:
                                if i == 1:
                                    spine_data[peer1][tmp_key] = tmp_list
                                else:
                                    spine_data[peer2][tmp_key] = tmp_list
                            elif common_values:
                                spine_data[peer1][tmp_key] = tmp_list
                                spine_data[peer2][tmp_key] = tmp_list
                            continue
                        row[i].value = row[i].value.strip()

                if network_details:
                    spine_data[peer1][tmp_key] = row[1].value
                    spine_data[peer2][tmp_key] = row[2].value
                elif common_values:
                    spine_data[peer1][tmp_key] = row[1].value
                    spine_data[peer2][tmp_key] = row[1].value

        excel_data['hosts'].update(spine_data)

        return excel_data

    def _convert_leaf_sheet(self, excel_data, leaf_sheet, architecture):
        """
        Converts the Leaf Network sheet for the Spine/Leaf VRD excel workbook

        :param excel_data (dict): a dictionary containing processed information
        from excel file
        :param leaf_sheet (openpyxl sheet): openpyxl Leaf Network sheet from
        workbook
        :param architecture (str): a string describing which DCN architecture
        is being processed

        Returns:
            excel_data: a dictionary containing newly processed Leaf data
            previously and processed information from excel file
        """
        data_row = -1
        leaf_data = {}
        network_details = False
        vsx_pair_info = False
        common_values = False
        leaf_hostname_row = None
        leaf_vsx_pair_row = None
        vsx_pair_info_dict = {}

        if architecture == 'ARCH3_EBGP':
            arch_value_dict = self.ARCH3_EBGP_EXCEL_DICT.copy()
            arch_value_dict.update(self.ARCH3_EXCEL_DICT)
        elif architecture == 'ARCH3_IBGP':
            arch_value_dict = self.ARCH3_IBGP_EXCEL_DICT.copy()
            arch_value_dict.update(self.ARCH3_EXCEL_DICT)

        for row in leaf_sheet.iter_rows():
            if row[0].row == 1:
                continue
            # Retreiving host data from section starting with 'Network Details'
            elif row[0].value == 'Network Details':
                network_details = True
                data_row = row[0].row + 1

                tmp_row = leaf_sheet[row[0].row + 1]
                if tmp_row[0].value == 'Device Hostname':
                    leaf_hostname_row = tmp_row[0].row
                    # Retrieves hostnames from next row
                    for tmp in tmp_row[1:]:
                        if tmp.value:
                            leaf_data[tmp.value] = {'role': 'leaf'}

            elif row[0].value == 'Leaf Common Values':
                common_values = True
                data_row = row[0].row + 2
            elif row[0].value == 'Leaf VSX Pair Information':
                vsx_pair_info = True
                data_row = row[0].row + 2

                tmp_row = leaf_sheet[row[0].row + 1]
                if tmp_row[0].value == 'Leaf VSX Pair':
                    leaf_vsx_pair_row = tmp_row[0].row
                    # Retrieves hostnames from next row
                    for tmp in tmp_row[1:]:
                        if tmp.value:
                            vsx_pair_info_dict[tmp.value.lower()] = {}

            elif not row[0].value and (network_details or vsx_pair_info or common_values):
                if network_details:
                    network_details = False
                elif vsx_pair_info:
                    vsx_pair_info = False
                    for device, device_data in leaf_data.items():
                        if 'vsx_pair' in device_data.keys():
                            leaf_data[device].update(vsx_pair_info_dict[device_data['vsx_pair']])
                elif common_values:
                    common_values = False
                data_row = -1
            # Process data for Leaf Common Values
            elif data_row <= row[0].row and row[0].value and common_values:
                tmp_key = row[0].value
                tmp_value = row[1].value

                if tmp_value and tmp_key:
                    # Ignore any cell where value is empty or n/a
                    if tmp_value in ['n/a', 'N/A', None, ""]:
                        continue
                    elif self._is_string(tmp_value):
                        if ',' in tmp_value:
                            tmp_var = tmp_value.split(',')
                            tmp_var = list(filter(None, tmp_var))
                            tmp_list = [x.strip() for x in tmp_var]

                            for device in leaf_data.keys():
                                leaf_data[device][tmp_key] = tmp_list

                        # Strip any extra white space
                        tmp_value = tmp_value.strip()

                    if tmp_key in arch_value_dict.keys():
                        tmp_key = arch_value_dict[tmp_key]
                    elif tmp_key in self.COMMON_EXCEL_DICT.keys():
                        tmp_key = self.COMMON_EXCEL_DICT[tmp_key]
                    elif tmp_key in self.VSX_EXCEL_DICT.keys():
                        tmp_key = self.VSX_EXCEL_DICT[tmp_key]

                    if tmp_key in ['server_vlans', 'trunk_vlans', 'vsx_isl_ports']:
                            if not isinstance(tmp_value, list):
                                tmp_value = [tmp_value]

                    if tmp_key == 'vsx_pair':
                        tmp_value = tmp_value.lower()

                    for device in leaf_data.keys():
                        leaf_data[device][tmp_key] = tmp_value

            # Process data for each peer under Network Details or Leaf VSX Pair information
            elif data_row <= row[0].row and row[0].value and (network_details or vsx_pair_info):
                tmp_key = row[0].value
                if tmp_key in arch_value_dict.keys():
                    tmp_key = arch_value_dict[tmp_key]
                elif tmp_key in self.COMMON_EXCEL_DICT.keys():
                    tmp_key = self.COMMON_EXCEL_DICT[tmp_key]
                elif tmp_key in self.VSX_EXCEL_DICT.keys():
                    tmp_key = self.VSX_EXCEL_DICT[tmp_key]

                # Convert matching cells into list values
                for tmp in row:
                    tmp_value = tmp.value
                    # Ignore any cell where value is empty or n/a
                    if tmp_value in ['n/a', 'N/A', None, ""]:
                        continue
                    if tmp.column is 1:
                        continue
                    if tmp_value:
                        if network_details:
                            column_key = leaf_sheet[leaf_hostname_row][tmp.column-1].value
                        elif vsx_pair_info:
                            column_key = leaf_sheet[leaf_vsx_pair_row][tmp.column-1].value.lower()
                            tmp_key = tmp_key.lower()
                        if self._is_string(tmp_value):
                            if ',' in tmp_value:
                                tmp_var = tmp_value.split(',')
                                tmp_var = list(filter(None, tmp_var))
                                tmp_list = [x.strip() for x in tmp_var]
                                if column_key and network_details:
                                    leaf_data[column_key][tmp_key] = tmp_list
                                elif column_key and vsx_pair_info:
                                    vsx_pair_info_dict[column_key][tmp_key] = tmp_list
                                continue

                            tmp_value = tmp_value.strip()

                        if tmp_key in ['server_vlans', 'vsx_isl_ports']:
                            if not isinstance(tmp_value, list):
                                tmp_value = [tmp_value]

                        if tmp_key == 'vsx_pair':
                            tmp_value = tmp_value.lower()

                        if column_key and network_details:
                            leaf_data[column_key][tmp_key] = tmp_value
                        elif column_key and vsx_pair_info:
                            vsx_pair_info_dict[column_key][tmp_key] = tmp_value

        for leaf in leaf_data.keys():
            excel_data['leaf'].append(leaf)

        excel_data['hosts'].update(leaf_data)

        return excel_data

    def _get_structured_inventory(self):
        """
        Processes excel workbook and determines DCN Architecture being
        definied

        Returns:
            excel_data (dict): a dictionary containing processed information
            from excel file
        """
        # Initialize a dict
        excel_data = {}
        excel_data['hosts'] = {}

        path = self.inv_dir+"/"+self.inv_file

        # Read the EXCEL and add it to the dictionary
        wb = load_workbook(filename=path, data_only=True)

        if all(item in wb.sheetnames for item in self.ARCH1_EXCEL_SHEETS):
            excel_data['architecture'] = 'ARCH1'
            excel_data['roles'] = ['dc_tor']
            excel_data['dc_tor'] = []
            excel_data = self._convert_network_sheet(excel_data, wb[self.ARCH1_EXCEL_SHEETS[0]], 'ARCH1')
            excel_data = self._convert_core_sheet(excel_data, wb[self.ARCH1_EXCEL_SHEETS[1]], 'ARCH1')
            excel_data = self._generate_loopback(excel_data, 'ARCH1')

        elif all(item in wb.sheetnames for item in self.ARCH2_EXCEL_SHEETS):
            excel_data['architecture'] = 'ARCH2'
            excel_data['roles'] = ['core', 'access']
            excel_data['core'] = []
            excel_data['access'] = []
            excel_data = self._convert_network_sheet(excel_data, wb[self.ARCH2_EXCEL_SHEETS[0]], 'ARCH2')
            excel_data = self._convert_core_sheet(excel_data, wb[self.ARCH2_EXCEL_SHEETS[1]], 'ARCH2')
            excel_data = self._convert_access_sheet(excel_data, wb[self.ARCH2_EXCEL_SHEETS[2]])
            excel_data = self._generate_loopback(excel_data, 'ARCH2')

        elif all(item in wb.sheetnames for item in self.ARCH3_EXCEL_SHEETS):
            excel_data['roles'] = ['spine', 'leaf']
            excel_data['spine'] = []
            excel_data['leaf'] = []
            excel_data = self._convert_network_sheet(excel_data, wb[self.ARCH3_EXCEL_SHEETS[0]], 'ARCH3')

            # Use eBGP as default Layer3 Fabric if none specified
            if 'architecture' not in excel_data.keys():
                excel_data['architecture'] = "ARCH3_EBGP"

            excel_data = self._convert_spine_sheet(excel_data, wb[self.ARCH3_EXCEL_SHEETS[1]], excel_data['architecture'])
            excel_data = self._convert_leaf_sheet(excel_data, wb[self.ARCH3_EXCEL_SHEETS[2]], excel_data['architecture'])
            excel_data = self._generate_loopback(excel_data, excel_data['architecture'])
            # Generate fabric data such as eBGP IP addressing, loopback addressing,
            # and interfaces for spine/leaf topology
            excel_data = self._define_fabric(excel_data)
        else:
            sys.exit('Error in Excel Sheet - Check Workbook Sheetnames')

        return excel_data

    def _define_fabric(self, excel_data):
        """
        Generates L3 fabric addressing, interfaces, and loopback addresses
        for a eBGP and iBGP fabric
        :param excel_data: resulting dictionary from _get_structured_inventory
        :return: dictionary of values for fabric, keys correspond to device
        hostname
        """
        # WIP Need to add logic to check if fabric is eBGP or iBGP
        if 'fabric_subnet' in excel_data['network'].keys():
            fabric_network = excel_data['network']['fabric_subnet']
        else:
            return excel_data

        if '/' not in fabric_network:
            fabric_network = fabric_network + "/24"

        try:
            fabric_network = IPNetwork(fabric_network)

        except AddrFormatError:
            exit("Error with Fabric Addresses value : {tmp_val}"
                 "".format(tmp_val=fabric_network))

        vsx_pairs = {}
        # Define lists of hosts which are either spines or leafs
        for leaf in excel_data['leaf']:
            host_data = excel_data['hosts'][leaf]
            if host_data['vsx_pair'] not in vsx_pairs.keys():
                if 'IBGP' in excel_data['architecture']:
                    vsx_pairs[host_data['vsx_pair']] = {'leafs': [leaf]}
                elif 'EBGP' in excel_data['architecture']:
                    vsx_pairs[host_data['vsx_pair']] = {'leafs': [leaf], 'vsx_pair_asn': host_data['vsx_pair_asn']}
            else:
                vsx_pair_data = vsx_pairs[host_data['vsx_pair']]
                vsx_pair_data['leafs'].append(leaf)

        fabric_dict = {}
        spines = sorted(excel_data['spine'])


        fabric_subnets = list(fabric_network.subnet(31))
        leaf_subnet_counter = 0
        spine_L3fabric_ips = {}

        spine_vsx_pairs = []
        spine_loopbacks = []

        # Define spine IP addressing for each leaf/vsx_pair downlink
        for spine in spines:
            fabric_dict[spine] = {}
            tmp_counter = leaf_subnet_counter

            for vsx_pair in sorted(vsx_pairs.keys()):
                fabric_dict[spine][vsx_pair] = {}
                tmp_vsx_pair_dictionary = []

                if vsx_pair not in spine_L3fabric_ips.keys():
                    spine_L3fabric_ips[vsx_pair] = []

                vsx_pair_downlinks = sorted(
                    excel_data['hosts'][spine]["vsx_pair_downlinks"][vsx_pair]["vsx_pair_downlink_interfaces"])

                for interface in vsx_pair_downlinks:
                    spine_ip = str(fabric_subnets[tmp_counter].ip + 1)
                    leaf_ip = str(fabric_subnets[tmp_counter].ip)
                    if 'IBGP' in excel_data['architecture']:
                        tmp_vsx_pair_dictionary.append(
                        {'vsx_pair_downlink_int': interface,
                            'vsx_pair_downlink_ip': spine_ip,
                            'vsx_pair_uplink_ip': leaf_ip})
                    elif 'EBGP' in excel_data['architecture']:
                        tmp_vsx_pair_dictionary.append(
                        {'vsx_pair_downlink_int': interface,
                            'vsx_pair_downlink_ip': spine_ip,
                            'vsx_pair_uplink_ip': leaf_ip,
                            'vsx_pair_asn': vsx_pairs[vsx_pair]['vsx_pair_asn']})

                    spine_L3fabric_ips[vsx_pair].append(spine_ip)

                    tmp_counter += 2

                fabric_dict[spine][vsx_pair] = tmp_vsx_pair_dictionary
            leaf_subnet_counter += 1
            spine_loopbacks.append(excel_data['hosts'][spine]['loopback0_ip'])

        leaf_subnet_counter = 0

        # Define Leaf IP addressing for each Spine Uplink
        for vsx_pair in sorted(vsx_pairs.keys()):
            leafs = sorted(vsx_pairs[vsx_pair]['leafs'])
            vsx_pair_loopbacks = []
            tmp_vsx_pair_asn = 0

            for leaf in leafs:
                if 'EBGP' in excel_data['architecture']:
                    tmp_vsx_pair_asn = excel_data['hosts'][leaf]['vsx_pair_asn']
                excel_data['hosts'][leaf]['spine_L3fabric_ips'] = spine_L3fabric_ips[vsx_pair]
                for interface in sorted(excel_data['hosts'][leaf]['spine_uplink_ports']):
                    leaf_ip = str(fabric_subnets[leaf_subnet_counter].ip)
                    if 'vsx_pair_L3fabric_ips' not in excel_data['hosts'][leaf].keys():
                        excel_data['hosts'][leaf]['vsx_pair_L3fabric_ips'] = []
                    excel_data['hosts'][leaf]['vsx_pair_L3fabric_ips'].append({'interface': interface,
                                        'ipv4': leaf_ip })

                    leaf_subnet_counter += 1
                vsx_pair_loopbacks.append(excel_data['hosts'][leaf]['loopback0_ip'])
                excel_data['hosts'][leaf]['spine_loopback0_ips'] = spine_loopbacks
            if 'EBGP' in excel_data['architecture']:
                spine_vsx_pairs.append({'vsx_pair_asn': tmp_vsx_pair_asn,
                                    'loopback_ips': vsx_pair_loopbacks})
            elif 'IBGP' in excel_data['architecture']:
                spine_vsx_pairs.append({'loopback_ips': vsx_pair_loopbacks})

        for spine in excel_data['spine']:
            excel_data['hosts'][spine]['vsx_pair_downlinks'] = fabric_dict[spine]
            excel_data['hosts'][spine]['vsx_pairs'] = spine_vsx_pairs

        return excel_data

    def _populate(self):
        '''Return the hosts and groups'''
        self.excel_data = self._get_structured_inventory()

        # Define Variables used for creating Ansible groups
        roles = self.excel_data['roles']
        vsx_pairs = []
        fabric_data = None
        # Gather all possible VSX groups as defined in worksheet in excel
        for hostname, data in self.excel_data['hosts'].items():
            if 'vsx_pair' in data.keys():
                vsx_pairs.append(data['vsx_pair'])

        for vsx_pair in vsx_pairs:
            self.inventory.add_group(vsx_pair)

        for function in roles:
            self.inventory.add_group(function)

        # Each device will be in the group aoscx_switches
        self.inventory.add_group('aoscx_switches')

        for hostname,host_data in self.excel_data['hosts'].items():
            self.inventory.add_host(host=hostname, group=host_data['role'])
            self.inventory.add_host(host=hostname, group='aoscx_switches')

            for data_key, data_val in host_data.items():
                self.inventory.set_variable(hostname, data_key, data_val)

            for data_key, data_val in self.excel_data['network'].items():
                self.inventory.set_variable(hostname, data_key, data_val)

            # Static variables that must always be set for our aoscx collection
            self.inventory.set_variable(hostname, 'ansible_network_os', 'arubanetworks.aoscx.aoscx')
            self.inventory.set_variable(hostname, 'ansible_connection', 'arubanetworks.aoscx.aoscx')
            self.inventory.set_variable(hostname, 'ansible_httpapi_use_ssl', True)

    def verify_file(self, path):
        '''Return true/false if this is a
        valid file for this plugin to consume
        '''
        valid = False
        if super(InventoryModule, self).verify_file(path):
            #base class verifies that file exists
            #and is readable by current user
            valid = True
            pass
        return valid

    def parse(self, inventory, loader, path, cache):
       '''Return dynamic inventory from source '''
       super(InventoryModule, self).parse(inventory, loader, path, cache)
       # Read the inventory YAML file
       self._read_config_data(path)
       try:
           # Store the options from the YAML file
           self.plugin = self.get_option('plugin')
           self.inv_dir = self.get_option('path_to_inventory')
           self.inv_file = self.get_option('excel_file')
       except Exception as e:
           raise AnsibleParserError(
               'All correct options required: {}'.format(e))
       # Call our internal helper to populate the dynamic inventory
       self._populate()
